from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl

app = Flask(__name__)
CORS(app)

nome = "asdqwe"
contratante = "123456"
json_data = {
    "lista1": [
        {"itens": "x", "DSCServico": "descriçaoramon aaaaaa", "VCUnid": 1, "VCQuant": 4, "VCVU": 7, "VCValorTotal": 1},
        {"itens": "y", "DSCServico": "descriçao aaaaaaaa", "VCUnid": 6, "VCQuant": 9, "VCVU": 6, "VCValorTotal": 34},
        {"itens": "z", "DSCServico": "descriçao aaaaaaaaaaaa", "VCUnid": 3, "VCQuant": 6, "VCVU": 12, "VCValorTotal": 9},
    ]
}

insert_row_index = 13

wb = openpyxl.load_workbook("BM Template - 01 UNT.xlsx")
sheet = wb.active  # Select the active sheet

for user in json_data['lista1']:
    sheet.insert_rows(insert_row_index, amount=1)

    sheet.cell(row=insert_row_index, column=1, value=user['itens'])
    sheet.merge_cells(start_row=insert_row_index, start_column=1, end_row=insert_row_index, end_column=2)
    sheet.cell(row=insert_row_index, column=3, value=user['DSCServico'])
    sheet.merge_cells(start_row=insert_row_index, start_column=3, end_row=insert_row_index, end_column=17)
    sheet.cell(row=insert_row_index, column=18, value=user['VCUnid'])
    sheet.cell(row=insert_row_index, column=21, value=user['VCQuant'])
    sheet.cell(row=insert_row_index, column=25, value=user['VCVU'])
    sheet.cell(row=insert_row_index, column=30, value=user['VCValorTotal'])

    insert_row_index += 1

wb.save("BM Template - 01 UNT_NEW.xlsx")


def replace_values_in_excel():
    replaceCells = {"[Nome]": nome, "[Contratante]": contratante}

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in replaceCells.keys():
                    cell.value = replaceCells.get(cell.value)


@app.route('/api/replace-excel-values', methods=['POST'])
def replace_excel_values():
    try:
        replace_values_in_excel()
        return jsonify({"status": "success", "message": "Sucesso, arquivo criado"})
    except Exception as error_mensage:
        return jsonify({"status": "error", "message": str(error_mensage)})


if __name__ == '__main__':
    app.run(debug=True)
